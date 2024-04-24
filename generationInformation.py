import random
import string


import pandas as pd
from openpyxl import Workbook

def gerar_nome():
    letras = string.ascii_lowercase
    nome = ''.join(random.choice(letras) for _ in range(8))
    return nome.capitalize()

#aqui vc esta declando o array e dizendo que ele tem 1000 posicoes
nome = ["" for x in range(1000)]
idade = ["" for x in range(1000)]
profissao = ["" for x in range(1000)]
lazer = ["" for x in range(1000)]
filiacao = ["" for x in range(1000)]


#aqui o for inicia em 0 e coloca no array (nome) na posicao 0 ate 1000 atraves da variavel dfe contriole x
for x in range(1000):
    nome[x] = gerar_nome()
    idade[x] = gerar_nome()
    profissao[x] = gerar_nome()
    lazer[x] = gerar_nome()
    filiacao[x] = gerar_nome()


wb = Workbook()
ws = wb.active


ws['A1'] = "Nome"
ws['B1'] = "Idade"
ws['C1'] = "profissao"
ws['D1'] = "Lazer"
ws['E1'] = "Filiacao"

dados = {
    'Nome': nome,
    'Idade': idade,
    'profissao': profissao,
    'Lazer': lazer,
    'Filiacao': filiacao

}
df = pd.DataFrame(dados)
df.to_excel('exemplo.xlsx', index=False)






