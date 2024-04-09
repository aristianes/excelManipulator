from openpyxl import Workbook
import random
import string

# Criar um objeto Workbook
wb = Workbook()

# Ativar a planilha ativa (por padrão, a primeira planilha é criada)
ws = wb.active

# Adicionar os dados nas posições desejadas
ws['A1'] = "Nome"
ws['B1'] = "Idade"
ws['C1'] = "Sexo"
ws['D1'] = "Endereco"


# Salvar a planilha
wb.save("exemplo.xlsx")




def gerar_nome():
    letras = string.ascii_lowercase
    nome = ''.join(random.choice(letras) for _ in range(8))
    return nome.capitalize()

with open('nome.txt','w') as arquivo:
    for _ in range(1000):
        nome = gerar_nome()
        arquivo.write(nome + '\n')

